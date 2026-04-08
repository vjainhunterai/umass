"""
Payment Terms Delta Manager
============================
A generic, reusable utility to:
  1. Load a master payment-terms file (Excel)
  2. Load incremental / source-system payment terms (text, CSV, or Excel)
  3. Parse every term into a canonical tuple (discount%, early_days, net_days)
  4. Compute the delta (new terms not yet in master)
  5. Append the delta to the master and produce a delta report

Usage
-----
    python payment_terms_delta_manager.py \
        --master  all_payment_terms.xlsx \
        --incremental umass_data_payment_terms.txt \
        --output  updated_master.xlsx \
        --report  delta_report.xlsx
"""

import argparse
import os
import re
from collections import defaultdict
from typing import Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ─────────────────────────────────────────────
# 1. PARSING ENGINE
# ─────────────────────────────────────────────

# Canonical immediate-payment keywords
IMMEDIATE_KEYWORDS = {
    "immediate", "due immediately", "due immediatly",
    "due on receipt", "pay immediately", "pay on receipt",
    "payable immediately", "payable upon receipt",
    "pay now", "fully prepaid", "prepayment",
    "net 0", "net0", "now",
}


def parse_payment_term(raw: str) -> tuple:
    """
    Parse any free-text payment term into a canonical tuple:
        (discount_pct: float | None,
         early_days:   int   | None,
         net_days:     int   | None)

    Handles formats like:
        "2% 10, Net 30"         -> (2.0, 10, 30)
        "2 % DISCOUNT 10, NET 30" -> (2.0, 10, 30)
        "Net 45"                -> (None, None, 45)
        "Immediate"             -> (None, None, 0)
        "1/2% 10 NET 30"       -> (0.5, 10, 30)
    """
    text = raw.strip()
    normalized = text.upper().strip()

    # --- Immediate / zero-day terms ---
    for kw in IMMEDIATE_KEYWORDS:
        if kw.upper() in normalized:
            return (None, None, 0)

    # --- Pure NET terms  (e.g. "Net 30", "NET30", "NET_30") ---
    net_only = re.match(
        r'^NET[_\s]*(\d+)\s*(DAYS?)?$', normalized
    )
    if net_only:
        return (None, None, int(net_only.group(1)))

    # --- Fraction discount  (e.g. "1/2% 10 NET 30") ---
    frac = re.match(
        r'^(\d+)/(\d+)%?\s*(?:DISC(?:OUNT)?\s*)?(\d+)\s*(?:DAYS?)?\s*'
        r'(?:,\s*)?NET\s*(\d+)',
        normalized
    )
    if frac:
        disc = float(frac.group(1)) / float(frac.group(2))
        return (disc, int(frac.group(3)), int(frac.group(4)))

    # --- Standard discount terms ---
    # Matches patterns like:
    #   "2% 10, Net 30"
    #   "2 % DISCOUNT 10, NET 30"
    #   ".5% 10 DAYS NET 30"
    #   "2%/10, NET 30"
    #   "2%10NET30"
    disc_pat = re.match(
        r'^\.?(\d+\.?\d*)\s*%\s*[/,]?\s*'           # discount
        r'(?:DISC(?:OUNT)?\s*(?:IN|IF PAID IN)?\s*)?'
        r'(\d+)\s*(?:DAYS?)?\s*'                      # early days
        r'[,\s]*(?:NET|N)\s*(\d+)',                    # net days
        normalized
    )
    if disc_pat:
        d = raw.strip()
        # Handle leading dot  (".5%" -> 0.5)
        disc_str = disc_pat.group(1)
        if d.startswith('.'):
            disc = float('0.' + disc_str)
        else:
            disc = float(disc_str)
        return (disc, int(disc_pat.group(2)), int(disc_pat.group(3)))

    # --- Verbose long-form  ("2 PERCENT DISCOUNT ... 10 DAYS ... 30 DAYS") ---
    verbose = re.search(
        r'(\d+\.?\d*)\s*(?:PERCENT|%)', normalized
    )
    if verbose:
        disc = float(verbose.group(1))
        days = re.findall(r'(\d+)\s*DAYS?', normalized)
        if len(days) >= 2:
            return (disc, int(days[0]), int(days[-1]))
        elif len(days) == 1:
            net_m = re.search(r'NET\s*(\d+)', normalized)
            if net_m:
                return (disc, int(days[0]), int(net_m.group(1)))

    # --- Fallback: try to find any NET number ---
    net_fb = re.search(r'NET\s*(\d+)', normalized)
    if net_fb:
        return (None, None, int(net_fb.group(1)))

    # --- Unparseable ---
    return (None, None, None)


# ─────────────────────────────────────────────
# 2. FILE LOADERS
# ─────────────────────────────────────────────

def load_master_excel(path: str) -> list[dict]:
    """
    Load master payment terms from an Excel file.
    Expects columns: key, description, discount, early_days, net_days
    Returns list of dicts.
    """
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    records = []
    for r in range(2, ws.max_row + 1):
        key = ws.cell(r, 1).value
        if key is None:
            break
        records.append({
            'key':       int(key),
            'desc':      str(ws.cell(r, 2).value or '').strip(),
            'discount':  ws.cell(r, 3).value,
            'early':     ws.cell(r, 4).value,
            'net':       ws.cell(r, 5).value,
        })
    return records


def load_incremental(path: str, column: Optional[str] = None) -> list[str]:
    """
    Load incremental payment terms from a text file, CSV, or Excel.
    Returns a list of raw term strings.
    """
    ext = os.path.splitext(path)[1].lower()

    if ext == '.txt':
        with open(path, 'r') as f:
            lines = [ln.strip() for ln in f if ln.strip()]
        # Skip header if it looks like one
        if lines and lines[0].lower().replace('_', ' ').startswith('default payment'):
            lines = lines[1:]
        return lines

    elif ext == '.csv':
        import csv
        with open(path, 'r') as f:
            reader = csv.DictReader(f)
            col = column or reader.fieldnames[0]
            return [row[col].strip() for row in reader if row[col].strip()]

    elif ext in ('.xlsx', '.xls'):
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        terms = []
        for r in range(2, ws.max_row + 1):
            val = ws.cell(r, 1).value
            if val:
                terms.append(str(val).strip())
        return terms

    else:
        raise ValueError(f"Unsupported file type: {ext}")


# ─────────────────────────────────────────────
# 3. DELTA COMPUTATION
# ─────────────────────────────────────────────

def normalize_key(discount, early, net) -> tuple:
    """Normalize a parsed tuple for comparison."""
    d = float(discount) if discount and discount != 0 else None
    e = int(early) if early and early != 0 else None
    try:
        n = int(net) if net is not None else None
    except (ValueError, TypeError):
        n = None
    return (d, e, n)


def compute_delta(master_records: list[dict],
                  incremental_terms: list[str]) -> dict:
    """
    Compare incremental terms against master.
    Returns dict with keys:
        matched          - list of (term, tuple, master_matches)
        new_terms         - list of (term, tuple) not in master
        master_only       - set of tuples in master but not incremental
        master_by_tuple   - lookup dict
    """
    # Build master lookup
    master_by_tuple = defaultdict(list)
    for rec in master_records:
        tup = normalize_key(rec['discount'], rec['early'], rec['net'])
        master_by_tuple[tup].append(rec)

    # Build master text lookup (exact description match)
    master_descs = {rec['desc'].upper() for rec in master_records}

    # Parse and classify incremental terms
    matched = []
    new_terms = []
    incr_tuples = set()

    for term in incremental_terms:
        parsed = parse_payment_term(term)
        canon = normalize_key(*parsed)
        incr_tuples.add(canon)

        if term.upper() in master_descs:
            matched.append((term, canon, master_by_tuple.get(canon, []),
                            'exact_text'))
        elif canon in master_by_tuple:
            matched.append((term, canon, master_by_tuple[canon],
                            'structural'))
        else:
            new_terms.append((term, canon))

    # Master-only tuples
    master_only = {t for t in master_by_tuple if t not in incr_tuples}

    return {
        'matched':        matched,
        'new_terms':      new_terms,
        'master_only':    master_only,
        'master_by_tuple': dict(master_by_tuple),
    }


# ─────────────────────────────────────────────
# 4. OUTPUT GENERATORS
# ─────────────────────────────────────────────

# Shared styles
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
HEADER_FONT   = Font(bold=True, color="FFFFFF", name="Arial", size=10)
DATA_FONT     = Font(name="Arial", size=10)
CENTER_ALIGN  = Alignment(horizontal="center", vertical="center")
WRAP_ALIGN    = Alignment(horizontal="center", vertical="center",
                          wrap_text=True)
YELLOW_FILL   = PatternFill("solid", fgColor="FFFFAA")
GREEN_FILL    = PatternFill("solid", fgColor="C8E6C9")
RED_FILL      = PatternFill("solid", fgColor="FFCDD2")
BLUE_FILL     = PatternFill("solid", fgColor="BBDEFB")


def _style_header(ws, row, headers, fill_color):
    fill = PatternFill("solid", fgColor=fill_color)
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row, c, h)
        cell.font = HEADER_FONT
        cell.fill = fill
        cell.alignment = WRAP_ALIGN
        cell.border = THIN_BORDER


def _style_row(ws, row, ncols, fill=None):
    for c in range(1, ncols + 1):
        cell = ws.cell(row, c)
        cell.font = DATA_FONT
        cell.border = THIN_BORDER
        cell.alignment = CENTER_ALIGN
        if fill:
            cell.fill = fill


def update_master(master_path: str,
                  new_terms: list[tuple],
                  output_path: str) -> int:
    """
    Append new terms to master Excel and save.
    Returns the number of terms added.
    """
    wb = openpyxl.load_workbook(master_path)
    ws = wb.active

    max_key = 0
    for r in range(2, ws.max_row + 1):
        k = ws.cell(r, 1).value
        if k is None:
            break
        max_key = max(max_key, int(k))

    next_key = max_key + 1
    for desc, (disc, early, net) in new_terms:
        r = ws.max_row + 1
        ws.cell(r, 1, next_key)
        ws.cell(r, 2, desc)
        ws.cell(r, 3, disc)
        ws.cell(r, 4, early)
        ws.cell(r, 5, net)
        _style_row(ws, r, 5, YELLOW_FILL)
        next_key += 1

    wb.save(output_path)
    return len(new_terms)


def generate_delta_report(delta: dict, output_path: str):
    """
    Create a multi-sheet delta report workbook.
    """
    wb = openpyxl.Workbook()

    # --- Sheet 1: Summary ---
    ws1 = wb.active
    ws1.title = "Summary"
    ws1['A1'] = "Payment Terms Delta Analysis"
    ws1['A1'].font = Font(bold=True, name="Arial", size=14, color="1565C0")
    ws1.merge_cells('A1:C1')

    _style_header(ws1, 3, ["Metric", "Count"], "424242")
    summary = [
        ("Incremental terms — matched (exact text)", 
         sum(1 for m in delta['matched'] if m[3] == 'exact_text')),
        ("Incremental terms — matched (structural)", 
         sum(1 for m in delta['matched'] if m[3] == 'structural')),
        ("Incremental terms — NEW (delta)",         
         len(delta['new_terms'])),
        ("Master-only combinations",                
         len(delta['master_only'])),
    ]
    for i, (label, val) in enumerate(summary, 4):
        ws1.cell(i, 1, label)
        ws1.cell(i, 2, val)
        _style_row(ws1, i, 2)
    ws1.column_dimensions['A'].width = 45
    ws1.column_dimensions['B'].width = 12

    # --- Sheet 2: New / Delta Terms ---
    ws2 = wb.create_sheet("Delta (New Terms)")
    _style_header(ws2, 1,
                  ["Term", "Discount %", "Early Days", "Net Days", "Action"],
                  "E65100")
    for i, (term, tup) in enumerate(delta['new_terms'], 2):
        ws2.cell(i, 1, term)
        ws2.cell(i, 2, tup[0])
        ws2.cell(i, 3, tup[1])
        ws2.cell(i, 4, tup[2])
        ws2.cell(i, 5, "ADD TO MASTER")
        _style_row(ws2, i, 5, RED_FILL)
    if not delta['new_terms']:
        ws2.cell(2, 1, "No new terms — all incremental terms exist in master")
        ws2.merge_cells('A2:E2')
        ws2['A2'].font = Font(name="Arial", size=11, italic=True,
                              color="2E7D32")
    ws2.column_dimensions['A'].width = 28
    for c in 'BCDE':
        ws2.column_dimensions[c].width = 14

    # --- Sheet 3: Matched Terms ---
    ws3 = wb.create_sheet("Matched Terms")
    _style_header(ws3, 1,
                  ["Incremental Term", "Discount %", "Early Days",
                   "Net Days", "Match Type", "Master Variations"],
                  "2E7D32")
    for i, (term, tup, recs, mtype) in enumerate(delta['matched'], 2):
        ws3.cell(i, 1, term)
        ws3.cell(i, 2, tup[0])
        ws3.cell(i, 3, tup[1])
        ws3.cell(i, 4, tup[2])
        ws3.cell(i, 5, mtype.replace('_', ' ').title())
        ws3.cell(i, 6, " | ".join(r['desc'] for r in recs[:5]))
        _style_row(ws3, i, 6, GREEN_FILL)
    ws3.column_dimensions['A'].width = 28
    ws3.column_dimensions['F'].width = 60
    for c in 'BCDE':
        ws3.column_dimensions[c].width = 14

    wb.save(output_path)


# ─────────────────────────────────────────────
# 5. CLI ENTRY POINT
# ─────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Payment Terms Delta Manager — "
                    "compare incremental terms against a master list"
    )
    parser.add_argument('--master', required=True,
                        help='Path to master payment terms Excel file')
    parser.add_argument('--incremental', required=True,
                        help='Path to incremental terms file '
                             '(.txt, .csv, or .xlsx)')
    parser.add_argument('--output', default='master_updated.xlsx',
                        help='Path for the updated master file')
    parser.add_argument('--report', default='delta_report.xlsx',
                        help='Path for the delta report')
    parser.add_argument('--column', default=None,
                        help='Column name for CSV incremental files')
    args = parser.parse_args()

    # Load
    print(f"Loading master:      {args.master}")
    master = load_master_excel(args.master)
    print(f"  → {len(master)} records")

    print(f"Loading incremental: {args.incremental}")
    incremental = load_incremental(args.incremental, args.column)
    print(f"  → {len(incremental)} terms")

    # Delta
    print("\nComputing delta...")
    delta = compute_delta(master, incremental)
    print(f"  Matched (exact text):  "
          f"{sum(1 for m in delta['matched'] if m[3] == 'exact_text')}")
    print(f"  Matched (structural):  "
          f"{sum(1 for m in delta['matched'] if m[3] == 'structural')}")
    print(f"  New terms (delta):     {len(delta['new_terms'])}")
    print(f"  Master-only combos:    {len(delta['master_only'])}")

    # Output
    if delta['new_terms']:
        added = update_master(args.master, delta['new_terms'], args.output)
        print(f"\n✓ Updated master saved: {args.output}  (+{added} terms)")
    else:
        print(f"\n✓ No new terms to add — master is up to date")

    generate_delta_report(delta, args.report)
    print(f"✓ Delta report saved:  {args.report}")


if __name__ == '__main__':
    main()
